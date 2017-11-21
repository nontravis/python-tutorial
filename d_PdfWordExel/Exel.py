# - Workbook
# - .xlsx file extension
# - Sheets / Worksheets
# - Columns (letters) & Rows (numbers)
# - Cell

#┌────────────────────┐
#│ NOTE: Reading exel │
#└────────────────────┘
import openpyxl
import os

workbook = openpyxl.load_workbook("d_PdfWordExel/example.xlsx")
type(workbook)
workbook.get_sheet_names()
sheet = workbook.get_sheet_by_name("Sheet1")
type(sheet)

cell = sheet["A1"]
cell.value
str(cell.value)

sheet["B1"].value
B = 2
sheet.cell(column=B, row=1).value  # same

sheet["C1"].value


#┌────────────────────┐
#│ NOTE: Ediging exel │
#└────────────────────┘
wb = openpyxl.Workbook()
wb.get_sheet_names()
sheet = wb.get_sheet_by_name("Sheet")

sheet["A1"].value == None
sheet["A1"] = "Hello"
sheet["B1"] = "World"
wb.save("d_PdfWordExel/example2.xlsx")

sheet2 = wb.create_sheet()
wb.get_sheet_names()
sheet2.title = "My new sheet name"
wb.save("d_PdfWordExel/example2.xlsx")
wb.create_sheet(index=0, title="My other sheet")
wb.save("d_PdfWordExel/example2.xlsx")
